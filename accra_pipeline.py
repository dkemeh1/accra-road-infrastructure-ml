# ============================================================
# MASTER STEPWISE PIPELINE (PUBLISH-SAFE v3.1): Accra Road Surface Mapping
# THREE SENSOR COMBINATIONS + STEP 8 (Compare validation & QC outputs)
#
# EXPERIMENTS (each writes to its own folder):
#   1) L9       : L8 (2015) + L8 (2024/2025)
#   2) S1       : S1 (2015) + S1 (2024/2025)
#   3) S2       : S2 (2015) + S2 (2024)
# Folder root:
#   C:\Users\dkeme\OneDrive\Desktop\ARIEL UNIVERSITY\MY PHD\DATA 2
#
# Steps:
#   STEP 1  - Extract UNIQUE surface tag values
#   STEP 2  - Create training labels (paved/unpaved) from OSM "surface" tag
#   STEP 3  - Segment ALL roads + create TRAIN segments
#   STEP 4  - Extract satellite features (EE download, resume-safe)
#   STEP 4B - Merge parts into one CSV per year
#   STEP 5  - Train + VALIDATE (Spatial CV)  [PUBLISH-SAFE]:
#              - NO leakage: imputation + clipping are fit on TRAIN fold only
#              - OOF probabilities built using fold-safe preprocessing
#              - TWO tuned thresholds on OOF:
#                   (A) best F1
#                   (B) best BALANCED ACCURACY
#              - Guardrail: reject thresholds that predict almost one class
#              - IMPORTANT FIX (v3.1):
#                   If guardrail rejects all thresholds, DO NOT accept collapse.
#                   Fall back to fixed threshold=0.50 and record "GUARDRAIL_FALLBACK".
#   STEP 6  - Predict ALL + QC (NOT accuracy)  [CONSISTENT PREPROCESSING]:
#              - Uses training medians + training clip bounds saved in model bundle
#              - Uses chosen threshold mode (BALACC default) CONSISTENTLY across experiments/years
#   STEP 7  - Change detection outputs (nearest-neighbor matching within distance)
#   STEP 8  - Compare experiments (reads Step 5/6/7 outputs)
# ============================================================

import os
from xgboost import XGBClassifier
from lightgbm import LGBMClassifier
import math
import time
import glob
import datetime
import requests
import xlsxwriter
import numpy as np
import pandas as pd
import geopandas as gpd
from shapely.geometry import LineString, MultiLineString
from shapely.ops import linemerge
import ee
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import GroupKFold
from sklearn.metrics import (
    accuracy_score,
    f1_score,
    confusion_matrix,
    balanced_accuracy_score,
    precision_score,
    recall_score,
)
import joblib


# ============================================================
# 0) GLOBAL CONFIG
# ============================================================
BASE_DIR = r"C:\Users\dkeme\OneDrive\Desktop\ARIEL UNIVERSITY\MY PHD\DATA 2"

YEAR_A = "2015"
YEAR_B = "2024"  # change to "2025" when you have it

SNAPSHOT_DATE_A = f"{YEAR_A}-11-30"
SNAPSHOT_DATE_B = f"{YEAR_B}-12-31"

SNAP_A = os.path.join(BASE_DIR, f"snapshot_{YEAR_A}_{YEAR_A}-12-31.geojson")
SNAP_B = os.path.join(BASE_DIR, f"snapshot_{YEAR_B}_{YEAR_B}-12-31.geojson")

WINDOW_DAYS = 30


# CRS + segmentation
METRIC_CRS = "EPSG:32630"   # UTM 30N
SEG_LEN_M = 100.0

# Step 4 batching
BATCH_SIZE = 6000
SLEEP_BETWEEN = 1
MAX_RETRIES = 5

# Step 5 spatial split (tile grouping)
TILE_SIZE_M = 2000

# Step 7 change detection
MAX_MATCH_DIST_M = 30

# ============================================================
# Step 5/6 KNOBS
# ============================================================
FEATURE_CLIP_QLOW = 0.01
FEATURE_CLIP_QHIGH = 0.99
THRESH_GRID = np.round(np.arange(0.05, 0.96, 0.01), 2)

# CONSISTENT RULE ACROSS ALL YEARS/EXPERIMENTS (DO NOT CHANGE AFTER LOOKING):
# - Step 5 reports BOTH best-F1 and best-BALACC
# - Step 6 uses ONLY this selection
THRESH_SELECTION_FOR_PREDICTION = "BALACC"   # "BALACC" or "F1"

# Guardrail against class-collapse thresholds
MIN_POS_RATE = 0.05   # at least 5% predicted positives
MAX_POS_RATE = 0.95   # at most 95% predicted positives

MODEL_TYPES = ["RF", "XGB", "LGBM"]   # choose any subset

RF_PARAMS = dict(
    n_estimators=500,
    random_state=42,
    n_jobs=-1,
    class_weight="balanced_subsample",
    max_features=0.5,
    min_samples_leaf=1,
    min_samples_split=4
)

XGB_PARAMS = dict(
    n_estimators=400,
    max_depth=6,
    learning_rate=0.05,
    subsample=0.8,
    colsample_bytree=0.8,
    objective="binary:logistic",
    eval_metric="logloss",
    n_jobs=-1,
    random_state=42
)

LGBM_PARAMS = dict(
    n_estimators=400,
    learning_rate=0.05,
    num_leaves=31,
    subsample=0.8,
    colsample_bytree=0.8,
    objective="binary",
    class_weight="balanced",
    n_jobs=-1,
    random_state=42
)

# ============================================================
# 0B) WHAT TO RUN
# ============================================================
RUN_EXPERIMENTS = [ "L8","S1","S2"]


RUN_STEP_1 = False
RUN_STEP_2 = False
RUN_STEP_3 = False
RUN_STEP_4 = False
RUN_STEP_4B = False
RUN_STEP_5_6 = False
RUN_STEP_7 = False
RUN_STEP_8 = True


# ============================================================
# 1) COMMON HELPERS
# ============================================================
def ensure_crs_wgs84(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    if gdf.crs is None:
        gdf = gdf.set_crs("EPSG:4326", allow_override=True)
    return gdf

def get_surface_series(gdf: gpd.GeoDataFrame) -> pd.Series:
    direct = gdf["surface"] if "surface" in gdf.columns else pd.Series([None]*len(gdf), index=gdf.index)
    if "tags" in gdf.columns:
        nested = gdf["tags"].apply(lambda x: x.get("surface") if isinstance(x, dict) else None)
    else:
        nested = pd.Series([None]*len(gdf), index=gdf.index)
    return direct.where(direct.notna(), nested)

def get_highway_series(gdf: gpd.GeoDataFrame) -> pd.Series:
    direct = gdf["highway"] if "highway" in gdf.columns else pd.Series([None]*len(gdf), index=gdf.index)
    if "tags" in gdf.columns:
        nested = gdf["tags"].apply(lambda x: x.get("highway") if isinstance(x, dict) else None)
    else:
        nested = pd.Series([None]*len(gdf), index=gdf.index)
    return direct.where(direct.notna(), nested)

def extract_osm_id(gdf: gpd.GeoDataFrame) -> pd.Series:
    for c in ["@id", "id", "osm_id", "osmid", "osmId", "@osmId"]:
        if c in gdf.columns:
            return gdf[c]
    if "tags" in gdf.columns:
        return gdf["tags"].apply(lambda x: x.get("@id") if isinstance(x, dict) else None)
    return pd.Series([None]*len(gdf), index=gdf.index)

def to_lines(geom):
    if geom is None:
        return None
    if isinstance(geom, LineString):
        return geom
    if isinstance(geom, MultiLineString):
        merged = linemerge(geom)
        if isinstance(merged, LineString):
            return merged
    return None

def split_line_into_segments(line: LineString, seg_len: float):
    length = line.length
    if length <= 0:
        return []
    n = max(1, int(math.ceil(length / seg_len)))
    segs = []
    for i in range(n):
        start_d = i * seg_len
        end_d = min((i + 1) * seg_len, length)
        if end_d - start_d < 1.0:
            continue
        steps = max(2, int((end_d - start_d) / 5))
        pts = [line.interpolate(start_d + (end_d - start_d) * (k / steps)) for k in range(steps + 1)]
        segs.append(LineString(pts))
    return segs


# ============================================================
# 2) EXPERIMENT PATHS
# ============================================================
def make_experiment_dirs(experiment_name: str) -> dict:
    EXP_ROOT = os.path.join(BASE_DIR, "experiments", experiment_name)
    OUT = {
        "EXP_ROOT": EXP_ROOT,
        "OUT_STEP1": os.path.join(EXP_ROOT, "surface_tags_only"),
        "OUT_STEP2": os.path.join(EXP_ROOT, "step2_labels"),
        "OUT_STEP3": os.path.join(EXP_ROOT, "step3_segments"),
        "OUT_STEP4_A": os.path.join(EXP_ROOT, f"step4_features_{YEAR_A}_6000"),
        "OUT_STEP4_B": os.path.join(EXP_ROOT, f"step4_features_{YEAR_B}_6000"),
        "OUT_STEP4_MERGED": os.path.join(EXP_ROOT, "step4_features"),
        "OUT_STEP5": os.path.join(EXP_ROOT, "step5_models"),
        "OUT_STEP6": os.path.join(EXP_ROOT, "step6_predictions"),
        "OUT_STEP7": os.path.join(EXP_ROOT, "step7_change_detection"),
        "OUT_STEP8": os.path.join(EXP_ROOT, "step8_compare"),
    }
    for d in OUT.values():
        os.makedirs(d, exist_ok=True)
    return OUT


# ============================================================
# STEP 1 — Unique surface tags
# ============================================================
def step1_extract_unique_surfaces(snapshot_path: str, year_label: str, OUT_STEP1: str):
    print("\n" + "="*70)
    print(f"STEP 1 — UNIQUE SURFACE TAG VALUES: {year_label}")
    print("="*70)

    gdf = ensure_crs_wgs84(gpd.read_file(snapshot_path))
    surface = get_surface_series(gdf)
    unique_vals = (
        surface.dropna().astype(str).str.strip()
        .replace("", pd.NA).dropna().unique().tolist()
    )
    unique_vals = sorted(set(unique_vals), key=lambda s: s.lower())

    out_txt = os.path.join(OUT_STEP1, f"surface_tags_{year_label}.txt")
    with open(out_txt, "w", encoding="utf-8") as f:
        for v in unique_vals:
            f.write(v + "\n")
    print(f"Saved: {out_txt}")


# ============================================================
# STEP 2 — Label training data (OSM weak supervision)
# ============================================================
PAVED_SET = {"asphalt", "concrete", "paved", "paving_stones", "sett", "concrete:plates"}
UNPAVED_SET = {"unpaved", "pebblestone", "grass", "metal", "wood", "rock",
               "gravel", "dirt", "earth", "ground", "sand", "mud", "fine_gravel", "compacted"}

def step2_label_snapshot(in_path: str, year_label: str, OUT_STEP2: str):
    print("\n" + "="*70)
    print(f"STEP 2 — TRAINING LABELS (PAVED=1 / UNPAVED=0): {year_label}")
    print("="*70)

    gdf = ensure_crs_wgs84(gpd.read_file(in_path))
    gdf["highway_extracted"] = get_highway_series(gdf)
    roads = gdf[gdf["highway_extracted"].notna()].copy()

    roads["surface_extracted"] = get_surface_series(roads).astype(str).str.strip().str.lower()
    roads.loc[roads["surface_extracted"].isin(["none", "nan", ""]), "surface_extracted"] = pd.NA
    roads = roads[roads["surface_extracted"].notna()].copy()

    roads["label"] = pd.NA
    roads.loc[roads["surface_extracted"].isin(PAVED_SET), "label"] = 1
    roads.loc[roads["surface_extracted"].isin(UNPAVED_SET), "label"] = 0

    train = roads[roads["label"].notna()].copy()
    train["label"] = train["label"].astype(int)
    train["year"] = int(year_label)

    out_gpkg = os.path.join(OUT_STEP2, f"train_{year_label}.gpkg")
    train.to_file(out_gpkg, layer="train", driver="GPKG")
    print(f"Saved: {out_gpkg} | rows={len(train):,}")


# ============================================================
# STEP 3 — Segment all roads + make train segments
# ============================================================
def step3_segment_all_roads(snapshot_path: str, year_label: str, OUT_STEP3: str):
    print("\n" + "="*70)
    print(f"STEP 3A — SEGMENT ALL ROADS (every {SEG_LEN_M}m): {year_label}")
    print("="*70)

    gdf = ensure_crs_wgs84(gpd.read_file(snapshot_path))
    gdf["highway"] = get_highway_series(gdf)
    gdf["surface"] = get_surface_series(gdf)
    gdf["osm_id"] = extract_osm_id(gdf)

    roads = gdf[gdf["highway"].notna()].copy()
    roads = roads[~roads.geometry.isna()].copy()
    roads = roads.to_crs(METRIC_CRS)
    roads["geometry"] = roads["geometry"].apply(to_lines)
    roads = roads[roads["geometry"].notna()].copy()

    rows = []
    for idx, r in roads.iterrows():
        segs = split_line_into_segments(r.geometry, SEG_LEN_M)
        for j, seg in enumerate(segs):
            rows.append({
                "seg_id": f"{year_label}_{idx}_{j}",
                "year": int(year_label),
                "osm_id": r.get("osm_id"),
                "highway": str(r.get("highway")),
                "surface": r.get("surface"),
                "geometry": seg
            })

    seg_gdf = gpd.GeoDataFrame(rows, crs=METRIC_CRS)
    out_all = os.path.join(OUT_STEP3, f"segments_{year_label}_all.gpkg")
    seg_gdf.to_file(out_all, layer="segments_all", driver="GPKG")
    print(f"Saved ALL segments: {len(seg_gdf):,} -> {out_all}")
    return out_all

def step3_make_train_segments(seg_all_gpkg: str, train_gpkg: str, year_label: str, OUT_STEP3: str):
    print("\n" + "="*70)
    print(f"STEP 3B — BUILD TRAIN SEGMENTS (intersects labeled roads): {year_label}")
    print("="*70)

    seg_all = gpd.read_file(seg_all_gpkg, layer="segments_all").to_crs(METRIC_CRS)
    train = ensure_crs_wgs84(gpd.read_file(train_gpkg, layer="train")).to_crs(METRIC_CRS)

    train_small = train[["label", "geometry"]].copy()
    joined = gpd.sjoin(seg_all, train_small, how="inner", predicate="intersects") \
                .drop(columns=["index_right"], errors="ignore")

    out_train = os.path.join(OUT_STEP3, f"segments_{year_label}_train.gpkg")
    joined.to_file(out_train, layer="segments_train", driver="GPKG")
    print(f"Saved TRAIN segments: {len(joined):,} -> {out_train}")
    return out_train


# ============================================================
# STEP 4 — Earth Engine init + download features
# ============================================================
def step4_init_ee():
    try:
        ee.Initialize()
    except Exception:
        ee.Authenticate()
        ee.Initialize()

def gdf_chunk_to_fc_centroids(gdf_wgs: gpd.GeoDataFrame) -> ee.FeatureCollection:
    feats = []
    for _, r in gdf_wgs.iterrows():
        geom = r.geometry
        if geom is None:
            continue
        c = geom.centroid
        ee_geom = ee.Geometry.Point([float(c.x), float(c.y)])
        props = {
            "seg_id": str(r["seg_id"]),
            "year": int(r["year"]),
            "highway": "" if pd.isna(r.get("highway", None)) else str(r.get("highway")),
            "surface": "" if pd.isna(r.get("surface", None)) else str(r.get("surface")),
        }
        feats.append(ee.Feature(ee_geom, props))
    return ee.FeatureCollection(feats)

def download_url_to_file(url: str, out_path: str):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            with requests.get(url, stream=True, timeout=300) as r:
                r.raise_for_status()
                with open(out_path, "wb") as f:
                    for chunk in r.iter_content(chunk_size=1024 * 1024):
                        if chunk:
                            f.write(chunk)
            return
        except Exception as e:
            print(f"Download failed (attempt {attempt}/{MAX_RETRIES}): {e}")
            time.sleep(2 * attempt)
    raise RuntimeError(f"Failed to download after {MAX_RETRIES} retries: {out_path}")

def bbox_aoi_from_segments(segments_gpkg: str) -> ee.Geometry:
    seg = gpd.read_file(segments_gpkg, layer="segments_all")
    seg = ensure_crs_wgs84(seg).to_crs("EPSG:4326")
    minx, miny, maxx, maxy = seg.total_bounds
    return ee.Geometry.Rectangle([float(minx), float(miny), float(maxx), float(maxy)])

# -------------------------------
# STEP 4 COMPOSITES
# -------------------------------
def composite_landsat8_sr_window(aoi, snapshot_date):
    start = ee.Date(snapshot_date).advance(-WINDOW_DAYS, "day")
    end   = ee.Date(snapshot_date).advance( WINDOW_DAYS, "day")

    col = (ee.ImageCollection("LANDSAT/LC08/C02/T1_L2")
           .filterBounds(aoi)
           .filterDate(start, end))

    img = col.median()   # <- KEY CHANGE

    sr = img.select(
        ["SR_B2", "SR_B3", "SR_B4", "SR_B5", "SR_B6", "SR_B7"]
    ).multiply(2.75e-05).add(-0.2)

    sr = sr.rename(["BLUE", "GREEN", "RED", "NIR", "SWIR1", "SWIR2"])

    ndvi = sr.normalizedDifference(["NIR", "RED"]).rename("NDVI")
    ndbi = sr.normalizedDifference(["SWIR1", "NIR"]).rename("NDBI")
    bright = sr.select(["RED", "NIR", "SWIR1"]).reduce(ee.Reducer.mean()).rename("BRIGHT")

    return sr.addBands([ndvi, ndbi, bright]).toFloat()


def composite_sentinel2_sr_window(aoi, snapshot_date):
    start = ee.Date(snapshot_date).advance(-WINDOW_DAYS, "day")
    end   = ee.Date(snapshot_date).advance( WINDOW_DAYS, "day")

    col = (ee.ImageCollection("COPERNICUS/S2_SR_HARMONIZED")
           .filterBounds(aoi)
           .filterDate(start, end)
           .filter(ee.Filter.lt("CLOUDY_PIXEL_PERCENTAGE", 40)))

    img = col.median()

    sr = img.select(
        ["B2", "B3", "B4", "B8", "B11", "B12"]
    ).divide(10000.0)

    sr = sr.rename(["BLUE", "GREEN", "RED", "NIR", "SWIR1", "SWIR2"])

    ndvi = sr.normalizedDifference(["NIR", "RED"]).rename("NDVI")
    ndbi = sr.normalizedDifference(["SWIR1", "NIR"]).rename("NDBI")
    bright = sr.select(["RED", "NIR", "SWIR1"]).reduce(ee.Reducer.mean()).rename("BRIGHT")

    return sr.addBands([ndvi, ndbi, bright]).toFloat()

def composite_sentinel1_sar_window(aoi, snapshot_date):
    start = ee.Date(snapshot_date).advance(-WINDOW_DAYS, "day")
    end   = ee.Date(snapshot_date).advance( WINDOW_DAYS, "day")

    col = (ee.ImageCollection("COPERNICUS/S1_GRD")
           .filterBounds(aoi)
           .filterDate(start, end)
           .filter(ee.Filter.eq("instrumentMode", "IW"))
           .filter(ee.Filter.listContains("transmitterReceiverPolarisation", "VV"))
           .filter(ee.Filter.listContains("transmitterReceiverPolarisation", "VH")))

    img = col.median()

    vv = ee.Image(10).multiply(img.select("VV").max(1e-6).log10()).rename("VV_db")
    vh = ee.Image(10).multiply(img.select("VH").max(1e-6).log10()).rename("VH_db")

    return ee.Image.cat([
        vv,
        vh,
        vv.subtract(vh).rename("VVminusVH"),
        vv.divide(vh).rename("VVdivVH")
    ]).toFloat()

def step4_extract_and_download(segments_gpkg: str, year_label: str, out_dir: str, composite_image, scale: int):
    print("\n" + "="*70)
    print(f"STEP 4 — FEATURE DOWNLOAD (batch={BATCH_SIZE}, resume): {year_label}")
    print("="*70)

    seg = gpd.read_file(segments_gpkg, layer="segments_all")
    seg = seg[[c for c in ["seg_id", "year", "highway", "surface", "geometry"] if c in seg.columns]].copy()
    seg = ensure_crs_wgs84(seg).to_crs("EPSG:4326")

    n = len(seg)
    parts = math.ceil(n / BATCH_SIZE)
    print(f"Total segments: {n:,} | Parts: {parts} | Output: {out_dir}")

    for p in range(parts):
        out_csv = os.path.join(out_dir, f"X_{year_label}_part{p+1:03d}.csv")
        if os.path.exists(out_csv) and os.path.getsize(out_csv) > 1000:
            print(f"Skip existing {p+1:03d}/{parts:03d}")
            continue

        chunk = seg.iloc[p*BATCH_SIZE: min((p+1)*BATCH_SIZE, n)].copy()
        fc = gdf_chunk_to_fc_centroids(chunk)

        sampled = composite_image.sampleRegions(
            collection=fc,
            properties=["seg_id", "year", "highway", "surface"],
            scale=scale,
            geometries=False
        )
        url = sampled.getDownloadURL(filetype="CSV")
        print(f"Downloading {p+1:03d}/{parts:03d} -> {out_csv}")
        download_url_to_file(url, out_csv)
        time.sleep(SLEEP_BETWEEN)

    print(f"DONE: {year_label} features saved -> {out_dir}")


# ============================================================
# STEP 4B — Merge parts
# ============================================================
def step4b_merge_parts(folder, prefix, out_csv):
    print("\n" + "="*70)
    print(f"STEP 4B — MERGE PARTS INTO ONE CSV: {prefix}")
    print("="*70)

    files = sorted(glob.glob(os.path.join(folder, f"X_{prefix}_part*.csv")))
    if not files:
        raise FileNotFoundError(f"No part files found in {folder} for {prefix}")

    valid = []
    skipped = 0

    for f in files:
        if os.path.getsize(f) < 50:
            print(f"⚠️ Skipped empty file: {os.path.basename(f)}")
            skipped += 1
            continue
        try:
            df = pd.read_csv(f)
            if df.empty:
                print(f"⚠️ Skipped empty dataframe: {os.path.basename(f)}")
                skipped += 1
                continue
            valid.append(df)
        except Exception as e:
            print(f"⚠️ Skipped unreadable file: {os.path.basename(f)} ({e})")
            skipped += 1

    if not valid:
        raise RuntimeError(f"All CSV parts were empty for {prefix}")

    out = pd.concat(valid, ignore_index=True)

    if "seg_id" in out.columns:
        out = out.drop_duplicates(subset=["seg_id"])

    out.to_csv(out_csv, index=False)
    print(f"✅ Merged {len(valid)} files (skipped {skipped}) → {out_csv}")

def make_model(model_type: str):
    model_type = model_type.upper()

    if model_type == "RF":
        return RandomForestClassifier(**RF_PARAMS)

    elif model_type == "XGB":
        return XGBClassifier(**XGB_PARAMS)

    elif model_type == "LGBM":
        return LGBMClassifier(**LGBM_PARAMS)

    else:
        raise ValueError(f"Unknown model_type: {model_type}")


# ============================================================
# STEP 5 — Train + VALIDATE (Spatial CV) [PUBLISH-SAFE: NO LEAKAGE]
# ============================================================
def add_tile_group(gdf):
    cent = gdf.geometry.centroid
    gdf["tx"] = (cent.x // TILE_SIZE_M).astype(int)
    gdf["ty"] = (cent.y // TILE_SIZE_M).astype(int)
    gdf["tile_id"] = gdf["tx"].astype(str) + "_" + gdf["ty"].astype(str)
    return gdf

def load_features(csv_path):
    df = pd.read_csv(csv_path)
    keep_non_num = {"seg_id", "year", "highway", "surface"}
    feat_cols = [c for c in df.columns if c not in keep_non_num and pd.api.types.is_numeric_dtype(df[c])]
    return df, feat_cols

def _compute_clip_bounds(X: pd.DataFrame, qlow=FEATURE_CLIP_QLOW, qhigh=FEATURE_CLIP_QHIGH) -> dict:
    bounds = {}
    for c in X.columns:
        s = X[c].dropna()
        if len(s) == 0:
            continue
        lo = float(s.quantile(qlow))
        hi = float(s.quantile(qhigh))
        if lo > hi:
            lo, hi = hi, lo
        bounds[c] = (lo, hi)
    return bounds

def _clip_with_bounds(X: pd.DataFrame, bounds: dict) -> pd.DataFrame:
    Xc = X.copy()
    for c, (lo, hi) in bounds.items():
        if c in Xc.columns:
            Xc[c] = Xc[c].clip(lower=lo, upper=hi)
    return Xc

def _passes_guardrail(pred: np.ndarray) -> bool:
    pos_rate = float(np.mean(pred == 1))
    return (pos_rate >= MIN_POS_RATE) and (pos_rate <= MAX_POS_RATE)

def _mcc_from_cm(cm: np.ndarray) -> float:
    tn, fp = float(cm[0, 0]), float(cm[0, 1])
    fn, tp = float(cm[1, 0]), float(cm[1, 1])
    denom = math.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))
    if denom == 0:
        return 0.0
    return (tp * tn - fp * fn) / denom

def _cm_row_normalized(cm: np.ndarray) -> np.ndarray:
    cm = cm.astype(float)
    row_sums = cm.sum(axis=1, keepdims=True)
    row_sums[row_sums == 0] = 1.0
    return cm / row_sums

def _metrics_from_pred(y_true: np.ndarray, pred: np.ndarray) -> dict:
    cm = confusion_matrix(y_true, pred, labels=[0, 1])
    tn, fp, fn, tp = int(cm[0, 0]), int(cm[0, 1]), int(cm[1, 0]), int(cm[1, 1])
    acc = float(accuracy_score(y_true, pred))
    f1 = float(f1_score(y_true, pred, zero_division=0))
    balacc = float(balanced_accuracy_score(y_true, pred))
    prec = float(precision_score(y_true, pred, zero_division=0))
    rec = float(recall_score(y_true, pred, zero_division=0))  # TPR
    tnr = float(tn / (tn + fp)) if (tn + fp) > 0 else 0.0
    mcc = float(_mcc_from_cm(cm))
    pos_rate = float(np.mean(pred == 1))
    cm_norm = _cm_row_normalized(cm)
    return {
        "acc": acc, "f1": f1, "balacc": balacc, "precision": prec, "recall": rec,
        "tpr": rec, "tnr": tnr, "mcc": mcc, "pos_rate": pos_rate,
        "cm": cm, "cm_norm": cm_norm,
        "tn": tn, "fp": fp, "fn": fn, "tp": tp
    }

def _search_thresholds(y_true: np.ndarray, proba: np.ndarray, grid=THRESH_GRID) -> dict:
    """
    FIXED v3.1:
    - We DO NOT fall back to an unguarded search that can accept degenerate thresholds.
    - If guardrail rejects all thresholds, we fall back to fixed threshold=0.50 (and mark it).
    """
    best_f1_score = -1.0
    best_ba_score = -1.0
    best_f1_row = None
    best_ba_row = None

    for t in grid:
        pred = (proba >= t).astype(int)
        if not _passes_guardrail(pred):
            continue
        m = _metrics_from_pred(y_true, pred)
        row = {"t": float(t), **m, "guardrail_fallback": False}
        if m["f1"] > best_f1_score:
            best_f1_score = m["f1"]
            best_f1_row = row
        if m["balacc"] > best_ba_score:
            best_ba_score = m["balacc"]
            best_ba_row = row

    if best_f1_row is None or best_ba_row is None:
        # Guardrail rejected everything -> fixed 0.50 fallback
        t = 0.50
        pred = (proba >= t).astype(int)
        m = _metrics_from_pred(y_true, pred)
        row = {"t": float(t), **m, "guardrail_fallback": True}
        # Use same row for both (still transparent + consistent)
        best_f1_row = row
        best_ba_row = row

    return {"F1": {"row": best_f1_row}, "BALACC": {"row": best_ba_row}}

def _print_threshold_report(year_label: str, tag: str, row: dict):
    cm = row["cm"]
    cmn = row["cm_norm"]
    fb = bool(row.get("guardrail_fallback", False))
    extra = "  [GUARDRAIL_FALLBACK_TO_0.50]" if fb else ""
    print("\n" + "-"*70)
    print(f"[{year_label}] OOF BEST {tag}: threshold={row['t']:.2f}{extra}")
    print(f"[{year_label}]   Accuracy:   {row['acc']:.3f}")
    print(f"[{year_label}]   F1:         {row['f1']:.3f}")
    print(f"[{year_label}]   BalAcc:     {row['balacc']:.3f}")
    print(f"[{year_label}]   Precision:  {row['precision']:.3f}")
    print(f"[{year_label}]   Recall/TPR: {row['tpr']:.3f}")
    print(f"[{year_label}]   TNR(spec):  {row['tnr']:.3f}")
    print(f"[{year_label}]   MCC:        {row['mcc']:.3f}")
    print(f"[{year_label}]   PosRate:    {row['pos_rate']:.3f}")
    print(f"[{year_label}]   Confusion matrix (rows=true 0/1, cols=pred 0/1):\n{cm}")
    print(f"[{year_label}]   Confusion matrix ROW-NORMALIZED:\n{np.round(cmn, 3)}")
    print("-"*70)

def step5_train_model_with_validation(year_label, train_seg_gpkg, feat_csv, model_out, eval_out_csv, model_type):
    print("\n" + "="*70)
    print(f"STEP 5 — TRAIN + VALIDATE (Spatial CV) [PUBLISH-SAFE v3.1]: {year_label}")
    print("="*70)

    seg_train = gpd.read_file(train_seg_gpkg, layer="segments_train").to_crs(METRIC_CRS)
    seg_train = add_tile_group(seg_train)

    Xdf, feat_cols = load_features(feat_csv)
    df = seg_train[["seg_id", "label", "tile_id"]].merge(Xdf, on="seg_id", how="inner").copy()

    if df.empty:
        raise RuntimeError(f"[{year_label}] No training rows after merge. Check seg_id match between GPKG and CSV.")

    y = df["label"].astype(int).values
    groups = df["tile_id"].astype(str).values

    Xraw = df[feat_cols].copy()

    gkf = GroupKFold(n_splits=5)
    oof_proba = np.full(shape=(len(Xraw),), fill_value=np.nan, dtype=float)

    fold_rows = []
    for fold, (tr, te) in enumerate(gkf.split(Xraw, y, groups), start=1):
        # --- Fold-safe preprocessing (NO LEAKAGE) ---
        X_tr = Xraw.iloc[tr].copy()
        X_te = Xraw.iloc[te].copy()

        med_tr = X_tr.median(numeric_only=True)
        X_tr = X_tr.fillna(med_tr)
        X_te = X_te.fillna(med_tr)

        clip_bounds_tr = _compute_clip_bounds(X_tr, FEATURE_CLIP_QLOW, FEATURE_CLIP_QHIGH)
        X_tr = _clip_with_bounds(X_tr, clip_bounds_tr)
        X_te = _clip_with_bounds(X_te, clip_bounds_tr)

        clf = make_model(model_type)
        clf.fit(X_tr, y[tr])

        proba = clf.predict_proba(X_te)[:, 1]
        oof_proba[te] = proba

        pred50 = (proba >= 0.50).astype(int)
        m50 = _metrics_from_pred(y[te], pred50)

        fold_rows.append({
            "year": int(year_label),
            "fold": fold,
            "accuracy@0.50": float(m50["acc"]),
            "f1@0.50": float(m50["f1"]),
            "balacc@0.50": float(m50["balacc"]),
            "mcc@0.50": float(m50["mcc"]),
            "tpr@0.50": float(m50["tpr"]),
            "tnr@0.50": float(m50["tnr"]),
        })
        print(f"Fold {fold}: acc@0.50={m50['acc']:.3f}  f1@0.50={m50['f1']:.3f}  balacc@0.50={m50['balacc']:.3f}")

    # --- OOF threshold tuning (guardrail-safe) ---
    mask = ~np.isnan(oof_proba)
    y_oof = y[mask]
    p_oof = oof_proba[mask]

    best = _search_thresholds(y_oof, p_oof, THRESH_GRID)
    row_f1 = best["F1"]["row"]
    row_ba = best["BALACC"]["row"]

    _print_threshold_report(year_label, "F1", row_f1)
    _print_threshold_report(year_label, "BALACC", row_ba)

    # Means across folds at fixed 0.50
    mean_acc_050 = float(np.mean([r["accuracy@0.50"] for r in fold_rows]))
    mean_f1_050 = float(np.mean([r["f1@0.50"] for r in fold_rows]))
    mean_balacc_050 = float(np.mean([r["balacc@0.50"] for r in fold_rows]))
    mean_mcc_050 = float(np.mean([r["mcc@0.50"] for r in fold_rows]))

    eval_df = pd.DataFrame(fold_rows)

    summary_mean = pd.DataFrame([{
        "year": int(year_label),
        "fold": "MEAN",
        "accuracy@0.50": mean_acc_050,
        "f1@0.50": mean_f1_050,
        "balacc@0.50": mean_balacc_050,
        "mcc@0.50": mean_mcc_050,
        "n_train_rows": int(len(df)),
        "n_features": int(len(feat_cols)),
        "note": "MEAN over folds at fixed threshold=0.50 (primary, conservative reporting)"
    }])

    def _summary(tag, row):
        cm = row["cm"]
        fb = bool(row.get("guardrail_fallback", False))
        return {
            "year": int(year_label),
            "fold": f"OOF_BEST_{tag}",
            "best_threshold": float(row["t"]),
            "accuracy@bestT": float(row["acc"]),
            "f1@bestT": float(row["f1"]),
            "balacc@bestT": float(row["balacc"]),
            "precision@bestT": float(row["precision"]),
            "recall_tpr@bestT": float(row["tpr"]),
            "tnr@bestT": float(row["tnr"]),
            "mcc@bestT": float(row["mcc"]),
            "pos_rate@bestT": float(row["pos_rate"]),
            "cm_tn": int(cm[0, 0]),
            "cm_fp": int(cm[0, 1]),
            "cm_fn": int(cm[1, 0]),
            "cm_tp": int(cm[1, 1]),
            "n_train_rows": int(len(df)),
            "n_features": int(len(feat_cols)),
            "guardrail_fallback": fb,
            "note": "OOF tuned threshold (from OOF predictions; not an external holdout). If guardrail_fallback=True, threshold forced to 0.50."
        }

    summary_f1 = pd.DataFrame([_summary("F1", row_f1)])
    summary_ba = pd.DataFrame([_summary("BALACC", row_ba)])

    out_eval = pd.concat([eval_df, summary_mean, summary_f1, summary_ba], ignore_index=True)
    out_eval.to_csv(eval_out_csv, index=False)
    print(f"Saved eval CSV: {eval_out_csv}")

    # --- Train FINAL model on ALL training data (consistent stats saved for Step 6) ---
    med_full = Xraw.median(numeric_only=True)
    X_full = Xraw.fillna(med_full)
    clip_bounds_full = _compute_clip_bounds(X_full, FEATURE_CLIP_QLOW, FEATURE_CLIP_QHIGH)
    X_full = _clip_with_bounds(X_full, clip_bounds_full)

    final = make_model(model_type)
    final.fit(X_full, y)

    # Choose which threshold Step 6 will use (CONSISTENT RULE)
    if THRESH_SELECTION_FOR_PREDICTION.upper() == "BALACC":
        chosen_t = float(row_ba["t"])
        chosen_tag = "BALACC"
    else:
        chosen_t = float(row_f1["t"])
        chosen_tag = "F1"

    joblib.dump({
        "model": final,
        "features": feat_cols,
        "train_medians": med_full.to_dict(),
        "clip_bounds": clip_bounds_full,
        "threshold_best_f1": float(row_f1["t"]),
        "threshold_best_balacc": float(row_ba["t"]),
        "threshold_used": float(chosen_t),
        "threshold_mode": chosen_tag,
        "guardrail_fallback_bestf1": bool(row_f1.get("guardrail_fallback", False)),
        "guardrail_fallback_bestba": bool(row_ba.get("guardrail_fallback", False)),
        "provenance": "Publish-safe v3.1: fold-safe CV preprocessing; Step6 uses ONE fixed rule (BALACC default) across all runs"
    }, model_out)

    print(f"Saved model bundle: {model_out}  (Step 6 will use {chosen_tag} threshold = {chosen_t:.2f})")


# ============================================================
# STEP 6 — Predict ALL + QC (NOT accuracy) [CONSISTENT WITH TRAINING]
# ============================================================
def _ensure_feature_columns(df: pd.DataFrame, feat_cols: list, fill_values: dict) -> pd.DataFrame:
    out = df.copy()
    for c in feat_cols:
        if c not in out.columns:
            out[c] = np.nan
        if c in fill_values:
            out[c] = out[c].fillna(fill_values[c])
    return out

def step6_predict_all_with_qc(year_label, seg_all_gpkg, feat_csv, model_path, out_gpkg, qc_out_csv):
    print("\n" + "="*70)
    print(f"STEP 6 — PREDICT ALL + QC [PUBLISH-SAFE v3.1]: {year_label}")
    print("="*70)

    seg_all = gpd.read_file(seg_all_gpkg, layer="segments_all").to_crs(METRIC_CRS)
    Xdf, _ = load_features(feat_csv)

    bundle = joblib.load(model_path)
    clf = bundle["model"]
    feat_cols = bundle["features"]
    clip_bounds = bundle.get("clip_bounds", {})
    train_medians = bundle.get("train_medians", {})

    threshold = float(bundle.get("threshold_used", 0.50))
    threshold_mode = str(bundle.get("threshold_mode", "0.50"))

    df = seg_all.merge(Xdf, on="seg_id", how="left")

    miss_feat = float(df[feat_cols].isna().mean().mean()) if all(c in df.columns for c in feat_cols) else np.nan

    df = _ensure_feature_columns(df, feat_cols, train_medians)
    Xmat = df[feat_cols].copy()
    Xmat = Xmat.fillna(0.0)

    if isinstance(clip_bounds, dict) and clip_bounds:
        Xmat = _clip_with_bounds(Xmat, clip_bounds)

    proba = clf.predict_proba(Xmat)[:, 1]
    pred = (proba >= threshold).astype(int)

    df["p_paved"] = proba
    df["pred_label"] = pred
    df["pred_surface"] = df["pred_label"].map({1: "paved", 0: "unpaved"})
    df["threshold_used"] = threshold
    df["threshold_mode"] = threshold_mode

    out = gpd.GeoDataFrame(df, crs=METRIC_CRS)
    out.to_file(out_gpkg, layer="predicted", driver="GPKG")
    print(f"Saved predictions: {out_gpkg}")

    qc = {
        "year": int(year_label),
        "threshold_used": float(threshold),
        "threshold_mode": threshold_mode,
        "n_segments": int(len(df)),
        "mean_missing_feature_rate": float(miss_feat) if pd.notna(miss_feat) else np.nan,
        "p_paved_mean": float(np.mean(proba)),
        "p_paved_std": float(np.std(proba)),
        "p_paved_min": float(np.min(proba)),
        "p_paved_max": float(np.max(proba)),
        "pct_pred_paved": float(np.mean(pred == 1) * 100.0),
        "pct_pred_unpaved": float(np.mean(pred == 0) * 100.0),
    }
    pd.DataFrame([qc]).to_csv(qc_out_csv, index=False)
    print(f"Saved QC CSV: {qc_out_csv}")


# ============================================================
# STEP 7 — Change detection (within SAME experiment)
# ============================================================
LAYER_BY_CLASS = {
    "Upgrade_unpaved_to_paved": "upgrade_unpaved_to_paved",
    "Downgrade_paved_to_unpaved": "downgrade_paved_to_unpaved",
    "Stable_paved": "stable_paved",
    "Stable_unpaved": "stable_unpaved",
    "Unmatched_or_unknown": "unmatched_or_unknown",
}

def step7_change_class(sA, sB):
    if sA == "unpaved" and sB == "paved":
        return "Upgrade_unpaved_to_paved"
    if sA == "paved" and sB == "unpaved":
        return "Downgrade_paved_to_unpaved"
    if sA == "paved" and sB == "paved":
        return "Stable_paved"
    if sA == "unpaved" and sB == "unpaved":
        return "Stable_unpaved"
    return "Unmatched_or_unknown"

def step7_run_change_detection(OUT_STEP6: str, OUT_STEP7: str):
    print("\n" + "="*70)
    print(f"STEP 7 — CHANGE DETECTION ({YEAR_A} → {YEAR_B})")
    print("="*70)

    pred_a = os.path.join(OUT_STEP6, f"segments_{YEAR_A}_predicted.gpkg")
    pred_b = os.path.join(OUT_STEP6, f"segments_{YEAR_B}_predicted.gpkg")

    gA = gpd.read_file(pred_a, layer="predicted").to_crs(METRIC_CRS)
    gB = gpd.read_file(pred_b, layer="predicted").to_crs(METRIC_CRS)

    gB_small = gB[["seg_id", "pred_surface", "p_paved", "geometry"]].copy().rename(columns={
        "seg_id": f"seg_id_{YEAR_B}",
        "pred_surface": f"surface_{YEAR_B}",
        "p_paved": f"p_paved_{YEAR_B}",
    })

    joined = gpd.sjoin_nearest(
        gA, gB_small,
        how="left",
        max_distance=MAX_MATCH_DIST_M,
        distance_col="match_dist_m"
    )

    joined = joined.rename(columns={
        "pred_surface": f"surface_{YEAR_A}",
        "p_paved": f"p_paved_{YEAR_A}",
    })

    joined["change_class"] = joined.apply(
        lambda r: step7_change_class(r.get(f"surface_{YEAR_A}"), r.get(f"surface_{YEAR_B}")),
        axis=1
    )
    joined["len_km"] = joined.geometry.length / 1000.0

    out_csv = os.path.join(OUT_STEP7, f"surface_change_summary_{YEAR_A}_{YEAR_B}.csv")
    summary = (joined.groupby("change_class")["len_km"].sum().reset_index().sort_values("len_km", ascending=False))
    summary["len_km"] = summary["len_km"].round(3)
    summary.to_csv(out_csv, index=False)

    out_gpkg = os.path.join(OUT_STEP7, f"surface_change_{YEAR_A}_{YEAR_B}.gpkg")
    joined.to_file(out_gpkg, layer="change_all", driver="GPKG")

    for cls, layer_name in LAYER_BY_CLASS.items():
        sub = joined[joined["change_class"] == cls].copy()
        if len(sub) > 0:
            sub.to_file(out_gpkg, layer=layer_name, driver="GPKG")

    print("Saved change GPKG:", out_gpkg)
    print("Saved summary CSV:", out_csv)
    print(summary.to_string(index=False))


# ============================================================
# STEP 8 — Compare experiments (NO penalties)
# ============================================================
def _read_row(eval_csv: str, fold_name: str) -> dict:
    if not os.path.exists(eval_csv):
        return {}
    df = pd.read_csv(eval_csv)
    row = df[df["fold"].astype(str).str.upper() == fold_name.upper()]
    if row.empty:
        return {}
    return row.iloc[0].to_dict()

def read_eval_for_report(eval_csv: str) -> dict:
    mean = _read_row(eval_csv, "MEAN")
    best_f1 = _read_row(eval_csv, "OOF_BEST_F1")
    best_ba = _read_row(eval_csv, "OOF_BEST_BALACC")

    out = {}
    if mean:
        out["acc_050"] = float(mean.get("accuracy@0.50", np.nan))
        out["f1_050"] = float(mean.get("f1@0.50", np.nan))
        out["balacc_050"] = float(mean.get("balacc@0.50", np.nan))
        out["mcc_050"] = float(mean.get("mcc@0.50", np.nan))

    def _pull(prefix, row):
        if not row:
            return
        out[f"{prefix}_t"] = float(row.get("best_threshold", np.nan))
        out[f"{prefix}_acc"] = float(row.get("accuracy@bestT", np.nan))
        out[f"{prefix}_f1"] = float(row.get("f1@bestT", np.nan))
        out[f"{prefix}_balacc"] = float(row.get("balacc@bestT", np.nan))
        out[f"{prefix}_mcc"] = float(row.get("mcc@bestT", np.nan))
        out[f"{prefix}_tpr"] = float(row.get("recall_tpr@bestT", np.nan))
        out[f"{prefix}_tnr"] = float(row.get("tnr@bestT", np.nan))
        out[f"{prefix}_pos_rate"] = float(row.get("pos_rate@bestT", np.nan))
        for k in ["cm_tn", "cm_fp", "cm_fn", "cm_tp"]:
            v = row.get(k, np.nan)
            out[f"{prefix}_{k}"] = int(float(v)) if pd.notna(v) else np.nan

    _pull("F1", best_f1)
    _pull("BALACC", best_ba)
    return out

def read_qc(qc_csv: str) -> dict:
    if not os.path.exists(qc_csv):
        return {}
    df = pd.read_csv(qc_csv)
    if df.empty:
        return {}
    return df.iloc[0].to_dict()

def read_change_summary(summary_csv: str) -> dict:
    if not os.path.exists(summary_csv):
        return {}
    df = pd.read_csv(summary_csv)
    if df.empty or "change_class" not in df.columns or "len_km" not in df.columns:
        return {}
    total = float(df["len_km"].sum())
    unmatched = float(df.loc[df["change_class"] == "Unmatched_or_unknown", "len_km"].sum()) \
        if "Unmatched_or_unknown" in set(df["change_class"]) else 0.0
    return {"unmatched_pct_step7": (unmatched / total * 100.0) if total > 0 else np.nan}

def safe_to_csv(df: pd.DataFrame, path: str) -> str:
    try:
        df.to_csv(path, index=False)
        return path
    except PermissionError:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.replace(".csv", f"_{ts}.csv")
        df.to_csv(alt, index=False)
        return alt

def step8_compare_and_report(experiment_names):
    rows = []

    for exp in experiment_names:
        exp_root = os.path.join(BASE_DIR, "experiments", exp)

        sensor, model_type = exp.split("_")

        eval_a = os.path.join(
            exp_root, "step5_models", f"eval_{model_type}_{YEAR_A}.csv"
        )
        eval_b = os.path.join(
            exp_root, "step5_models", f"eval_{model_type}_{YEAR_B}.csv"
        )

        qc_a = os.path.join(exp_root, "step6_predictions", f"qc_{YEAR_A}.csv")
        qc_b = os.path.join(exp_root, "step6_predictions", f"qc_{YEAR_B}.csv")

        chg = os.path.join(exp_root, "step7_change_detection",
                           f"surface_change_summary_{YEAR_A}_{YEAR_B}.csv")

        eA = read_eval_for_report(eval_a)
        eB = read_eval_for_report(eval_b)
        qA = read_qc(qc_a)
        qB = read_qc(qc_b)
        cS = read_change_summary(chg)

        rows.append({
            "experiment": exp,

            # 2015 best-F1
            "F1_2015_bestF1": eA.get("F1_f1", np.nan),
            "BALACC_2015_bestF1": eA.get("F1_balacc", np.nan),
            "MCC_2015_bestF1": eA.get("F1_mcc", np.nan),
            "TPR_2015_bestF1": eA.get("F1_tpr", np.nan),
            "TNR_2015_bestF1": eA.get("F1_tnr", np.nan),
            "T_2015_bestF1": eA.get("F1_t", np.nan),
            "cm2015_tn_bestF1": eA.get("F1_cm_tn", np.nan),
            "cm2015_fp_bestF1": eA.get("F1_cm_fp", np.nan),
            "cm2015_fn_bestF1": eA.get("F1_cm_fn", np.nan),
            "cm2015_tp_bestF1": eA.get("F1_cm_tp", np.nan),
            "posrate_2015_bestF1": eA.get("F1_pos_rate", np.nan),

            # 2015 best-BALACC
            "F1_2015_bestBA": eA.get("BALACC_f1", np.nan),
            "BALACC_2015_bestBA": eA.get("BALACC_balacc", np.nan),
            "MCC_2015_bestBA": eA.get("BALACC_mcc", np.nan),
            "TPR_2015_bestBA": eA.get("BALACC_tpr", np.nan),
            "TNR_2015_bestBA": eA.get("BALACC_tnr", np.nan),
            "T_2015_bestBA": eA.get("BALACC_t", np.nan),
            "cm2015_tn_bestBA": eA.get("BALACC_cm_tn", np.nan),
            "cm2015_fp_bestBA": eA.get("BALACC_cm_fp", np.nan),
            "cm2015_fn_bestBA": eA.get("BALACC_cm_fn", np.nan),
            "cm2015_tp_bestBA": eA.get("BALACC_cm_tp", np.nan),
            "posrate_2015_bestBA": eA.get("BALACC_pos_rate", np.nan),

            # 2024 best-F1
            "F1_2024_bestF1": eB.get("F1_f1", np.nan),
            "BALACC_2024_bestF1": eB.get("F1_balacc", np.nan),
            "MCC_2024_bestF1": eB.get("F1_mcc", np.nan),
            "TPR_2024_bestF1": eB.get("F1_tpr", np.nan),
            "TNR_2024_bestF1": eB.get("F1_tnr", np.nan),
            "T_2024_bestF1": eB.get("F1_t", np.nan),
            "cm2024_tn_bestF1": eB.get("F1_cm_tn", np.nan),
            "cm2024_fp_bestF1": eB.get("F1_cm_fp", np.nan),
            "cm2024_fn_bestF1": eB.get("F1_cm_fn", np.nan),
            "cm2024_tp_bestF1": eB.get("F1_cm_tp", np.nan),
            "posrate_2024_bestF1": eB.get("F1_pos_rate", np.nan),

            # 2024 best-BALACC
            "F1_2024_bestBA": eB.get("BALACC_f1", np.nan),
            "BALACC_2024_bestBA": eB.get("BALACC_balacc", np.nan),
            "MCC_2024_bestBA": eB.get("BALACC_mcc", np.nan),
            "TPR_2024_bestBA": eB.get("BALACC_tpr", np.nan),
            "TNR_2024_bestBA": eB.get("BALACC_tnr", np.nan),
            "T_2024_bestBA": eB.get("BALACC_t", np.nan),
            "cm2024_tn_bestBA": eB.get("BALACC_cm_tn", np.nan),
            "cm2024_fp_bestBA": eB.get("BALACC_cm_fp", np.nan),
            "cm2024_fn_bestBA": eB.get("BALACC_cm_fn", np.nan),
            "cm2024_tp_bestBA": eB.get("BALACC_cm_tp", np.nan),
            "posrate_2024_bestBA": eB.get("BALACC_pos_rate", np.nan),

            # fixed-0.50 mean (primary conservative)
            "F1_2015_mean@0.50": eA.get("f1_050", np.nan),
            "BALACC_2015_mean@0.50": eA.get("balacc_050", np.nan),
            "MCC_2015_mean@0.50": eA.get("mcc_050", np.nan),
            "F1_2024_mean@0.50": eB.get("f1_050", np.nan),
            "BALACC_2024_mean@0.50": eB.get("balacc_050", np.nan),
            "MCC_2024_mean@0.50": eB.get("mcc_050", np.nan),

            "missing_feat_rate_2015": qA.get("mean_missing_feature_rate", np.nan),
            "missing_feat_rate_2024": qB.get("mean_missing_feature_rate", np.nan),
            "unmatched_pct_step7": cS.get("unmatched_pct_step7", np.nan),
        })

    comp = pd.DataFrame(rows)

    out_dir = os.path.join(BASE_DIR, "experiments", "_COMPARISON_STEP8")
    os.makedirs(out_dir, exist_ok=True)

    out_csv = os.path.join(out_dir, f"comparison_step8_{YEAR_A}_{YEAR_B}.csv")
    saved_path = safe_to_csv(comp, out_csv)

    print("\n" + "="*120)
    print("STEP 8 — EXPERIMENT COMPARISON TABLE (NO penalties; includes MCC/TPR/TNR + CM fields)")
    print("="*120)
    print(comp.to_string(index=False))
    print(f"\nSaved comparison CSV: {saved_path}")


# ============================================================
# EXPERIMENT RUNNER
# ============================================================
def run_experiment(experiment_name: str):
    sensor, model_type = experiment_name.split("_")

    OUT = make_experiment_dirs(experiment_name)

    print("\n" + "#"*90)
    print(f"RUNNING EXPERIMENT: {experiment_name}")
    print("#"*90)

    if RUN_STEP_1:
        step1_extract_unique_surfaces(SNAP_A, YEAR_A, OUT["OUT_STEP1"])
        step1_extract_unique_surfaces(SNAP_B, YEAR_B, OUT["OUT_STEP1"])

    if RUN_STEP_2:
        step2_label_snapshot(SNAP_A, YEAR_A, OUT["OUT_STEP2"])
        step2_label_snapshot(SNAP_B, YEAR_B, OUT["OUT_STEP2"])

    if RUN_STEP_3:
        seg_all_a = step3_segment_all_roads(SNAP_A, YEAR_A, OUT["OUT_STEP3"])
        seg_all_b = step3_segment_all_roads(SNAP_B, YEAR_B, OUT["OUT_STEP3"])

        train_a = os.path.join(OUT["OUT_STEP2"], f"train_{YEAR_A}.gpkg")
        train_b = os.path.join(OUT["OUT_STEP2"], f"train_{YEAR_B}.gpkg")

        step3_make_train_segments(seg_all_a, train_a, YEAR_A, OUT["OUT_STEP3"])
        step3_make_train_segments(seg_all_b, train_b, YEAR_B, OUT["OUT_STEP3"])

    if RUN_STEP_4:
        step4_init_ee()

        seg_all_a = os.path.join(OUT["OUT_STEP3"], f"segments_{YEAR_A}_all.gpkg")
        seg_all_b = os.path.join(OUT["OUT_STEP3"], f"segments_{YEAR_B}_all.gpkg")

        aoi_a = bbox_aoi_from_segments(seg_all_a)
        aoi_b = bbox_aoi_from_segments(seg_all_b)



        if sensor == "L8":
            img_a = composite_landsat8_sr_window(aoi_a, SNAPSHOT_DATE_A)
            img_b = composite_landsat8_sr_window(aoi_b, SNAPSHOT_DATE_B)
            scale_a = scale_b = 30

        elif sensor == "S1":
            img_a = composite_sentinel1_sar_window(aoi_a, SNAPSHOT_DATE_A)
            img_b = composite_sentinel1_sar_window(aoi_b, SNAPSHOT_DATE_B)
            scale_a = scale_b = 10

        elif sensor == "S2":
            img_a = composite_sentinel2_sr_window(aoi_a, SNAPSHOT_DATE_A)
            img_b = composite_sentinel2_sr_window(aoi_b, SNAPSHOT_DATE_B)
            scale_a = scale_b = 10

        else:
            raise ValueError(f"Unknown sensor type: {sensor}")

        step4_extract_and_download(seg_all_a, YEAR_A, OUT["OUT_STEP4_A"], img_a, scale=scale_a)
        step4_extract_and_download(seg_all_b, YEAR_B, OUT["OUT_STEP4_B"], img_b, scale=scale_b)

    if RUN_STEP_4B:
        step4b_merge_parts(OUT["OUT_STEP4_A"], YEAR_A, os.path.join(OUT["OUT_STEP4_MERGED"], f"X_{YEAR_A}.csv"))
        step4b_merge_parts(OUT["OUT_STEP4_B"], YEAR_B, os.path.join(OUT["OUT_STEP4_MERGED"], f"X_{YEAR_B}.csv"))

    if RUN_STEP_5_6:
        seg_all_a = os.path.join(OUT["OUT_STEP3"], f"segments_{YEAR_A}_all.gpkg")
        seg_all_b = os.path.join(OUT["OUT_STEP3"], f"segments_{YEAR_B}_all.gpkg")
        seg_train_a = os.path.join(OUT["OUT_STEP3"], f"segments_{YEAR_A}_train.gpkg")
        seg_train_b = os.path.join(OUT["OUT_STEP3"], f"segments_{YEAR_B}_train.gpkg")

        X_a = os.path.join(OUT["OUT_STEP4_MERGED"], f"X_{YEAR_A}.csv")
        X_b = os.path.join(OUT["OUT_STEP4_MERGED"], f"X_{YEAR_B}.csv")

        for p in [seg_all_a, seg_all_b, seg_train_a, seg_train_b, X_a, X_b]:
            if not os.path.exists(p):
                raise FileNotFoundError(f"Missing required file for Step 5/6:\n{p}")

        model_a = os.path.join(OUT["OUT_STEP5"], f"model_{model_type}_{YEAR_A}.joblib")
        model_b = os.path.join(OUT["OUT_STEP5"], f"model_{model_type}_{YEAR_B}.joblib")

        eval_a = os.path.join(OUT["OUT_STEP5"], f"eval_{model_type}_{YEAR_A}.csv")
        eval_b = os.path.join(OUT["OUT_STEP5"], f"eval_{model_type}_{YEAR_B}.csv")

        step5_train_model_with_validation(
            YEAR_A, seg_train_a, X_a, model_a, eval_a, model_type
        )
        step5_train_model_with_validation(
            YEAR_B, seg_train_b, X_b, model_b, eval_b, model_type
        )

        step6_predict_all_with_qc(
            YEAR_A, seg_all_a, X_a, model_a,
            os.path.join(OUT["OUT_STEP6"], f"segments_{YEAR_A}_predicted.gpkg"),
            os.path.join(OUT["OUT_STEP6"], f"qc_{YEAR_A}.csv")
        )
        step6_predict_all_with_qc(
            YEAR_B, seg_all_b, X_b, model_b,
            os.path.join(OUT["OUT_STEP6"], f"segments_{YEAR_B}_predicted.gpkg"),
            os.path.join(OUT["OUT_STEP6"], f"qc_{YEAR_B}.csv")
        )

    if RUN_STEP_7:
        step7_run_change_detection(OUT["OUT_STEP6"], OUT["OUT_STEP7"])


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    for exp in RUN_EXPERIMENTS:
        for model_type in MODEL_TYPES:
            run_experiment(f"{exp}_{model_type}")

    if RUN_STEP_8:
        FULL_EXPERIMENTS = [
            f"{sensor}_{model}"
            for sensor in RUN_EXPERIMENTS
            for model in MODEL_TYPES
        ]

        step8_compare_and_report(FULL_EXPERIMENTS)

# ============================================================
# STEP 9 — CONFUSION MATRICES (ONE EXCEL, MULTIPLE WORKSHEETS)
# ============================================================

# ============================================================
# CONFIG
# ============================================================
BASE_DIR = r"C:\Users\dkeme\OneDrive\Desktop\ARIEL UNIVERSITY\MY PHD\DATA 2"
YEAR_A = "2015"
YEAR_B = "2024"

IN_CSV = os.path.join(
    BASE_DIR,
    "experiments",
    "_COMPARISON_STEP8",
    f"comparison_step8_{YEAR_A}_{YEAR_B}.csv"
)

OUT_XLSX = os.path.join(
    BASE_DIR,
    "experiments",
    "_COMPARISON_STEP8",
    "comparison_step8_FINAL_READABLE.xlsx"
)

# ============================================================
# LOAD
# ============================================================
df = pd.read_csv(IN_CSV)

# ============================================================
# HELPER TO EXTRACT LONG FORMAT
# ============================================================
def extract_rows(year, tag):
    rows = []
    for _, r in df.iterrows():
        exp = r["experiment"]
        suffix = f"{year}_{tag}"

        rows.append({
            "experiment": exp,
            "year": int(year),
            "threshold_type": "BEST_F1" if tag == "bestF1" else "BEST_BALACC",
            "threshold": r.get(f"T_{suffix}", np.nan),
            "accuracy": r.get(f"F1_{suffix}", np.nan),
            "f1": r.get(f"F1_{suffix}", np.nan),
            "balanced_accuracy": r.get(f"BALACC_{suffix}", np.nan),
            "mcc": r.get(f"MCC_{suffix}", np.nan),
            "tpr": r.get(f"TPR_{suffix}", np.nan),
            "tnr": r.get(f"TNR_{suffix}", np.nan),
            "tn": r.get(f"cm{year}_tn_{tag}", np.nan),
            "fp": r.get(f"cm{year}_fp_{tag}", np.nan),
            "fn": r.get(f"cm{year}_fn_{tag}", np.nan),
            "tp": r.get(f"cm{year}_tp_{tag}", np.nan),
        })
    return pd.DataFrame(rows)

# ============================================================
# BUILD DATAFRAMES
# ============================================================
dfs = {
    "2015_BEST_F1": extract_rows(YEAR_A, "bestF1"),
    "2015_BEST_BALACC": extract_rows(YEAR_A, "bestBA"),
    "2024_BEST_F1": extract_rows(YEAR_B, "bestF1"),
    "2024_BEST_BALACC": extract_rows(YEAR_B, "bestBA"),
}

summary_rows = []
for name, d in dfs.items():
    for _, r in d.iterrows():
        summary_rows.append({
            "experiment": r["experiment"],
            "year": r["year"],
            "threshold_type": r["threshold_type"],
            "accuracy": r["accuracy"],
            "balanced_accuracy": r["balanced_accuracy"],
            "mcc": r["mcc"],
            "tpr": r["tpr"],
            "tnr": r["tnr"],
        })

summary_df = pd.DataFrame(summary_rows)

# ============================================================
# WRITE EXCEL
# ============================================================
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    for sheet, d in dfs.items():
        d.to_excel(writer, sheet_name=sheet, index=False)
    summary_df.to_excel(writer, sheet_name="SUMMARY_CORE_METRICS", index=False)

# ============================================================
# POST-PROCESS EXCEL (FORMATTING + CONFUSION MATRICES)
# ============================================================
wb = load_workbook(OUT_XLSX)

for sheet_name in dfs.keys():
    ws = wb[sheet_name]

    ws["P1"] = "CONFUSION MATRIX (RAW)"
    ws["P1"].font = Font(bold=True)

    ws["P3"] = "TN"
    ws["Q3"] = "FP"
    ws["P4"] = "FN"
    ws["Q4"] = "TP"

    for i, row in enumerate(dfs[sheet_name].itertuples(), start=5):
        ws[f"P{i}"] = row.tn
        ws[f"Q{i}"] = row.fp
        ws[f"P{i+1}"] = row.fn
        ws[f"Q{i+1}"] = row.tp
        break  # only first experiment visualized per sheet (clean)

    rule = ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"
    )
    ws.conditional_formatting.add(f"P5:Q6", rule)

wb.save(OUT_XLSX)

print("✅ FINAL EXCEL CREATED:")
print(OUT_XLSX)
