import os
import warnings
import math
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import mannwhitneyu, ttest_ind
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

warnings.filterwarnings("ignore")

# ============================================================
# 0) SETTINGS
# ============================================================
PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = PROJECT_ROOT / "output_part3"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OUT_FILE = OUTPUT_DIR / "transport_exclusion_results.xlsx"

BLIND_FILE = PROJECT_ROOT / "Blindspots_final.csv"
NONBLIND_FILE = PROJECT_ROOT / "Non_blind_final.csv"
STABLE_FILE = PROJECT_ROOT / "Stable_unpaved_roads_fixed.csv"
DISTRICT_FILE = PROJECT_ROOT / "Blindspots_districts.csv"

# ============================================================
# 1) HELPER FUNCTIONS
# ============================================================
def safe_read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    return pd.read_csv(path)


def choose_first_existing(df: pd.DataFrame, candidates: list[str], required: bool = True):
    for col in candidates:
        if col in df.columns:
            return col
    if required:
        raise KeyError(f"None of these columns were found: {candidates}")
    return None


def cliffs_delta(x: pd.Series, y: pd.Series) -> float:
    x = pd.Series(x).dropna().values
    y = pd.Series(y).dropna().values

    if len(x) == 0 or len(y) == 0:
        return np.nan

    gt = 0
    lt = 0
    for xi in x:
        gt += np.sum(xi > y)
        lt += np.sum(xi < y)
    return (gt - lt) / (len(x) * len(y))


def valid_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").dropna()


def summarize_series(series: pd.Series) -> dict:
    s = valid_numeric(series)
    if len(s) == 0:
        return {
            "n": 0,
            "sum": np.nan,
            "mean": np.nan,
            "median": np.nan,
            "std": np.nan,
            "min": np.nan,
            "max": np.nan,
        }
    return {
        "n": int(len(s)),
        "sum": float(s.sum()),
        "mean": float(s.mean()),
        "median": float(s.median()),
        "std": float(s.std(ddof=1)) if len(s) > 1 else 0.0,
        "min": float(s.min()),
        "max": float(s.max()),
    }


def compare_groups(x: pd.Series, y: pd.Series, variable_name: str):
    x = valid_numeric(x)
    y = valid_numeric(y)

    if len(x) < 2 or len(y) < 2:
        return None

    result = {"variable": variable_name}

    u_stat, u_p = mannwhitneyu(x, y, alternative="two-sided")
    result["mannwhitney_u"] = float(u_stat)
    result["mannwhitney_p"] = float(u_p)

    t_stat, t_p = ttest_ind(x, y, equal_var=False, nan_policy="omit")
    result["welch_t"] = float(t_stat)
    result["welch_p"] = float(t_p)

    result["cliffs_delta"] = float(cliffs_delta(x, y))

    result["blind_n"] = int(len(x))
    result["nonblind_n"] = int(len(y))
    return result


def get_excel_col_letter(col_num: int) -> str:
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def auto_adjust_width(ws, max_width=40):
    for col_cells in ws.columns:
        length = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                length = max(length, len(value))
            except Exception:
                pass
        ws.column_dimensions[get_excel_col_letter(col_idx)].width = min(length + 2, max_width)


def style_sheet(ws, title=None):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)

    if title:
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="left")
        start_row = 3
    else:
        start_row = 1

    for row in ws.iter_rows(min_row=start_row, max_row=start_row):
        for cell in row:
            if cell.value is not None:
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

    auto_adjust_width(ws)


def nice_axis_max(max_value: float) -> float:
    if pd.isna(max_value) or max_value <= 0:
        return 1

    magnitude = 10 ** math.floor(math.log10(max_value))
    norm = max_value / magnitude

    if norm <= 1:
        nice = 1
    elif norm <= 2:
        nice = 2
    elif norm <= 5:
        nice = 5
    else:
        nice = 10

    return nice * magnitude


def add_bar_chart(
    ws,
    title,
    data_col,
    category_col,
    min_row,
    max_row,
    chart_cell,
    y_title="Value",
    x_title="Group",
    number_format='#,##0.00'
):
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.height = 8
    chart.width = 14

    data = Reference(ws, min_col=data_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=category_col, min_row=min_row + 1, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.dLbls.showPercent = False
    chart.dLbls.position = "outEnd"

    values = []
    for r in range(min_row + 1, max_row + 1):
        v = ws.cell(row=r, column=data_col).value
        try:
            if v is not None and not pd.isna(float(v)):
                values.append(float(v))
        except Exception:
            pass

    if values:
        max_val = max(values)
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = nice_axis_max(max_val * 1.15)

    chart.y_axis.numFmt = number_format
    ws.add_chart(chart, chart_cell)


# ============================================================
# 2) READ FILES
# ============================================================
blind_raw = safe_read_csv(BLIND_FILE)
nonblind_raw = safe_read_csv(NONBLIND_FILE)
stable_raw = safe_read_csv(STABLE_FILE)
district_raw = safe_read_csv(DISTRICT_FILE)

# ============================================================
# 3) CLEAN TABLES
# ============================================================
blind_zone_col = choose_first_existing(blind_raw, ["zone_id"])
blind_pop_sum_col = choose_first_existing(blind_raw, ["pop_sum"])
blind_pop_mean_col = choose_first_existing(blind_raw, ["pop_mean"])
blind_dist_col = choose_first_existing(blind_raw, ["HubDist"])

blind_clean = blind_raw[[blind_zone_col, blind_pop_sum_col, blind_pop_mean_col, blind_dist_col]].copy()
blind_clean.columns = ["zone_id", "pop_sum", "pop_mean", "dist_major"]

nonblind_zone_col = choose_first_existing(nonblind_raw, ["zone_id"])
nonblind_pop_sum_col = choose_first_existing(nonblind_raw, ["pop_sum"])
nonblind_pop_mean_col = choose_first_existing(nonblind_raw, ["pop_mean"])
nonblind_dist_col = choose_first_existing(nonblind_raw, ["HubDist"])

nonblind_clean = nonblind_raw[[nonblind_zone_col, nonblind_pop_sum_col, nonblind_pop_mean_col, nonblind_dist_col]].copy()
nonblind_clean.columns = ["zone_id", "pop_sum", "pop_mean", "dist_major"]

stable_highway_col = choose_first_existing(stable_raw, ["highway_x", "highway_y", "highway"], required=False)
stable_group_col = choose_first_existing(stable_raw, ["road_group"])

stable_keep = [stable_group_col]
if stable_highway_col:
    stable_keep.insert(0, stable_highway_col)

stable_clean = stable_raw[stable_keep].copy()
rename_map = {stable_group_col: "road_group"}
if stable_highway_col:
    rename_map[stable_highway_col] = "highway"
stable_clean = stable_clean.rename(columns=rename_map)
stable_clean["road_group"] = stable_clean["road_group"].astype(str).str.strip()

district_zone_col = choose_first_existing(district_raw, ["zone_id"])
district_name_col = choose_first_existing(district_raw, ["District", "Label"])
district_pop_sum_col = choose_first_existing(district_raw, ["pop_sum"], required=False)
district_pop_mean_col = choose_first_existing(district_raw, ["pop_mean"], required=False)

district_keep = [district_zone_col, district_name_col]
if district_pop_sum_col:
    district_keep.append(district_pop_sum_col)
if district_pop_mean_col:
    district_keep.append(district_pop_mean_col)

district_clean = district_raw[district_keep].copy()

rename_map = {
    district_zone_col: "zone_id",
    district_name_col: "district",
}
if district_pop_sum_col:
    rename_map[district_pop_sum_col] = "pop_sum_district"
if district_pop_mean_col:
    rename_map[district_pop_mean_col] = "pop_mean_district"

district_clean = district_clean.rename(columns=rename_map)

# ============================================================
# 4) RQ1 — POPULATION EXPOSURE
# ============================================================
blind_pop_sum_stats = summarize_series(blind_clean["pop_sum"])
blind_pop_mean_stats = summarize_series(blind_clean["pop_mean"])

nonblind_pop_sum_stats = summarize_series(nonblind_clean["pop_sum"])
nonblind_pop_mean_stats = summarize_series(nonblind_clean["pop_mean"])

rq1_summary = pd.DataFrame([
    {
        "Group": "Blindspots",
        "Zones": blind_pop_sum_stats["n"],
        "Total_Population": blind_pop_sum_stats["sum"],
        "Mean_Population_per_Zone": blind_pop_sum_stats["mean"],
        "Median_Population_per_Zone": blind_pop_sum_stats["median"],
        "Mean_Population_Intensity": blind_pop_mean_stats["mean"],
        "Median_Population_Intensity": blind_pop_mean_stats["median"],
    },
    {
        "Group": "Non-blind",
        "Zones": nonblind_pop_sum_stats["n"],
        "Total_Population": nonblind_pop_sum_stats["sum"],
        "Mean_Population_per_Zone": nonblind_pop_sum_stats["mean"],
        "Median_Population_per_Zone": nonblind_pop_sum_stats["median"],
        "Mean_Population_Intensity": nonblind_pop_mean_stats["mean"],
        "Median_Population_Intensity": nonblind_pop_mean_stats["median"],
    }
])

# ============================================================
# 5) RQ2 — ACCESSIBILITY
# ============================================================
blind_dist_stats = summarize_series(blind_clean["dist_major"])
nonblind_dist_stats = summarize_series(nonblind_clean["dist_major"])

rq2_summary = pd.DataFrame([
    {
        "Group": "Blindspots",
        "Zones": blind_dist_stats["n"],
        "Mean_Distance_to_Major_Road": blind_dist_stats["mean"],
        "Median_Distance_to_Major_Road": blind_dist_stats["median"],
        "Std_Distance": blind_dist_stats["std"],
        "Min_Distance": blind_dist_stats["min"],
        "Max_Distance": blind_dist_stats["max"],
    },
    {
        "Group": "Non-blind",
        "Zones": nonblind_dist_stats["n"],
        "Mean_Distance_to_Major_Road": nonblind_dist_stats["mean"],
        "Median_Distance_to_Major_Road": nonblind_dist_stats["median"],
        "Std_Distance": nonblind_dist_stats["std"],
        "Min_Distance": nonblind_dist_stats["min"],
        "Max_Distance": nonblind_dist_stats["max"],
    }
])

# ============================================================
# 6) RQ3 — ROAD HIERARCHY
# ============================================================
rq3_summary = (
    stable_clean["road_group"]
    .value_counts(dropna=False)
    .reset_index()
)
rq3_summary.columns = ["Road_Group", "Count"]
rq3_summary["Percentage"] = 100 * rq3_summary["Count"] / rq3_summary["Count"].sum()

# ============================================================
# 7) DISTRICT CONCENTRATION — KEEP ALL MATCHES
# ============================================================
blind_with_district_all = district_clean.merge(
    blind_clean[["zone_id", "pop_sum", "pop_mean", "dist_major"]],
    on="zone_id",
    how="left"
)

if "pop_sum_district" in blind_with_district_all.columns:
    blind_with_district_all["pop_sum_used"] = blind_with_district_all["pop_sum_district"].fillna(
        blind_with_district_all["pop_sum"]
    )
else:
    blind_with_district_all["pop_sum_used"] = blind_with_district_all["pop_sum"]

if "pop_mean_district" in blind_with_district_all.columns:
    blind_with_district_all["pop_mean_used"] = blind_with_district_all["pop_mean_district"].fillna(
        blind_with_district_all["pop_mean"]
    )
else:
    blind_with_district_all["pop_mean_used"] = blind_with_district_all["pop_mean"]

district_summary = (
    blind_with_district_all
    .groupby("district", dropna=False)
    .agg(
        blindspot_rows=("zone_id", "count"),
        unique_blindspots=("zone_id", "nunique"),
        total_population=("pop_sum_used", "sum"),
        mean_population=("pop_mean_used", "mean"),
        mean_distance_to_major_road=("dist_major", "mean"),
        median_distance_to_major_road=("dist_major", "median"),
    )
    .reset_index()
    .sort_values(by="total_population", ascending=False)
)

# ============================================================
# 8) STATISTICAL TESTS — ONLY VALID ONES
# ============================================================
stats_results = []

test_pop_mean = compare_groups(blind_clean["pop_mean"], nonblind_clean["pop_mean"], "pop_mean")
if test_pop_mean is not None:
    stats_results.append(test_pop_mean)

test_pop_sum = compare_groups(blind_clean["pop_sum"], nonblind_clean["pop_sum"], "pop_sum")
if test_pop_sum is not None:
    stats_results.append(test_pop_sum)

test_dist = compare_groups(blind_clean["dist_major"], nonblind_clean["dist_major"], "dist_major")
if test_dist is not None:
    stats_results.append(test_dist)

stats_df = pd.DataFrame(stats_results)

# ============================================================
# 9) WRITE EVERYTHING TO ONE EXCEL FILE
# ============================================================
with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
    readme_df = pd.DataFrame({
        "Item": [
            "RQ1",
            "RQ2",
            "RQ3",
            "District Summary",
            "Statistical Tests",
            "Important Note"
        ],
        "Description": [
            "Who is affected? Population exposure in blindspots vs non-blind areas.",
            "Do blindspots have poorer access to the main road network? Distance to major roads.",
            "Is underdevelopment concentrated on minor roads? Road hierarchy summary.",
            "District-level concentration using all district-blindspot matches exactly as shown in the data.",
            "Only valid statistical tests are included. If a test could not be computed, it is omitted.",
            "Non-blind area currently has very few rows, so some inferential tests may be unavailable."
        ]
    })
    readme_df.to_excel(writer, sheet_name="README", index=False)

    rq1_summary.to_excel(writer, sheet_name="RQ1_Population_Exposure", index=False, startrow=2)
    rq2_summary.to_excel(writer, sheet_name="RQ2_Accessibility", index=False, startrow=2)
    rq3_summary.to_excel(writer, sheet_name="RQ3_Road_Hierarchy", index=False, startrow=2)
    district_summary.to_excel(writer, sheet_name="District_Concentration", index=False, startrow=2)

    if not stats_df.empty:
        stats_df.to_excel(writer, sheet_name="Statistical_Tests", index=False, startrow=2)
    else:
        pd.DataFrame({
            "Message": ["No valid inferential tests could be computed with the current group sizes."]
        }).to_excel(writer, sheet_name="Statistical_Tests", index=False, startrow=2)

    blind_clean.to_excel(writer, sheet_name="Clean_Blindspots", index=False)
    nonblind_clean.to_excel(writer, sheet_name="Clean_NonBlind", index=False)
    stable_clean.to_excel(writer, sheet_name="Clean_StableRoads", index=False)
    district_clean.to_excel(writer, sheet_name="Clean_Districts", index=False)

# ============================================================
# 10) FORMAT WORKBOOK AND ADD EXCEL CHARTS
# ============================================================
wb = load_workbook(OUT_FILE)

for sheet_name, title in [
    ("README", "Transport Exclusion Analysis Workbook"),
    ("RQ1_Population_Exposure", "RQ1 — Who is affected? Population Exposure"),
    ("RQ2_Accessibility", "RQ2 — Accessibility to Major Roads"),
    ("RQ3_Road_Hierarchy", "RQ3 — Road Hierarchy of Stable Unpaved Roads"),
    ("District_Concentration", "District-Level Concentration"),
    ("Statistical_Tests", "Valid Statistical Tests Only"),
    ("Clean_Blindspots", "Clean Blindspots Table"),
    ("Clean_NonBlind", "Clean Non-Blind Table"),
    ("Clean_StableRoads", "Clean Stable Unpaved Roads Table"),
    ("Clean_Districts", "Clean District Table"),
]:
    ws = wb[sheet_name]
    style_sheet(ws, title=title)

ws1 = wb["RQ1_Population_Exposure"]
add_bar_chart(
    ws=ws1,
    title="Total Population by Group",
    data_col=3,
    category_col=1,
    min_row=3,
    max_row=5,
    chart_cell="J3",
    y_title="Population",
    x_title="Group",
    number_format='#,##0'
)

add_bar_chart(
    ws=ws1,
    title="Mean Population Intensity by Group",
    data_col=6,
    category_col=1,
    min_row=3,
    max_row=5,
    chart_cell="J20",
    y_title="Mean population intensity",
    x_title="Group",
    number_format='#,##0.00'
)

ws2 = wb["RQ2_Accessibility"]
add_bar_chart(
    ws=ws2,
    title="Mean Distance to Major Road by Group",
    data_col=3,
    category_col=1,
    min_row=3,
    max_row=5,
    chart_cell="J3",
    y_title="Distance (m)",
    x_title="Group",
    number_format='#,##0.00'
)

add_bar_chart(
    ws=ws2,
    title="Median Distance to Major Road by Group",
    data_col=4,
    category_col=1,
    min_row=3,
    max_row=5,
    chart_cell="J20",
    y_title="Distance (m)",
    x_title="Group",
    number_format='#,##0.00'
)

ws3 = wb["RQ3_Road_Hierarchy"]
max_row_rq3 = ws3.max_row
add_bar_chart(
    ws=ws3,
    title="Stable Unpaved Roads by Road Group (%)",
    data_col=3,
    category_col=1,
    min_row=3,
    max_row=max_row_rq3,
    chart_cell="F3",
    y_title="Percentage (%)",
    x_title="Road group",
    number_format='0.00'
)

ws4 = wb["District_Concentration"]
max_row_dist = ws4.max_row
if max_row_dist >= 4:
    add_bar_chart(
        ws=ws4,
        title="Total Population by District",
        data_col=4,
        category_col=1,
        min_row=3,
        max_row=max_row_dist,
        chart_cell="I3",
        y_title="Population",
        x_title="District",
        number_format='#,##0'
    )

wb.save(OUT_FILE)

print("=" * 70)
print("DONE")
print("=" * 70)
print(f"Excel workbook created successfully:\n{OUT_FILE}")
print(f"Output folder:\n{OUTPUT_DIR}")
print("\nSheets included:")
for s in wb.sheetnames:
    print(f"- {s}")